import re
import csv
import io
import json

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Min, Q
from django.urls import reverse
from django.views.decorators.http import require_POST as _require_POST
from django.utils.timezone import now as _now

from Email_validate_app.utils import get_user_id
from Email_validate_app.services.filter_utils import extract_filter_params, apply_search, apply_status, apply_date_range, timed_count
from Email_validate_app.services.filter_status import CONTACT_STATUSES


def _sync_campaign_list_counts(list_id):
    from Email_validate_app.models import CampaignList, CampaignEmail
    qs = CampaignEmail.objects.filter(list_id=list_id, deleted_at__isnull=True)
    CampaignList.objects.filter(id=list_id).update(
        total_count=qs.count(),
        subscribed_count=qs.filter(subscribed='subscribed').count(),
        neversubscribed_count=qs.filter(subscribed='never_subscribed').count(),
        unsubscribed_count=qs.filter(subscribed='unsubscribed').count(),
    )


def list_segment(request):
    if not request.session.get('logged_in'):
        messages.warning(request, "You need to login first.")
        return redirect(reverse('login'))

    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)

    if request.method == 'POST':
        import json as _json
        _body = _json.loads(request.body)
        list_name = _body.get('list_name', '').strip()
        tags = _body.get('tags', '').strip()
        if not list_name:
            return JsonResponse({'success': False, 'error': 'List name is required.'}, status=400)
        if CampaignList.objects.filter(user_id=user_id, list_name__iexact=list_name, deleted_at__isnull=True).exists():
            return JsonResponse({'success': False, 'error': f'A list named "{list_name}" already exists.'}, status=400)
        lst = CampaignList.objects.create(user_id=user_id, list_name=list_name, tags=tags)
        return JsonResponse({'success': True, 'id': lst.id})

    lists = CampaignList.objects.filter(user_id=user_id, deleted_at__isnull=True).order_by('-created_at')
    for lst in lists:
        lst.tags_list = [t.strip() for t in lst.tags.split(',') if t.strip()] if lst.tags else []
    return render(request, 'i_List_Segment.html', {
        'lists':        lists,
        'pf_statuses':  [('active', 'Active'), ('inactive', 'Inactive')],
    })


def list_rename(request, list_id):
    """Edit list name and/or status."""
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)
    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)
    try:
        body     = json.loads(request.body)
        new_name = (body.get('name') or '').strip()
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)
    if not new_name:
        return JsonResponse({'status': 'error', 'message': 'Name is required.'}, status=400)
    if len(new_name) > 255:
        return JsonResponse({'status': 'error', 'message': 'Name too long.'}, status=400)
    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found.'}, status=404)
    if CampaignList.objects.filter(user_id=user_id, list_name__iexact=new_name, deleted_at__isnull=True).exclude(id=list_id).exists():
        return JsonResponse({'status': 'error', 'message': f'A list named "{new_name}" already exists.'}, status=400)
    new_status = body.get('status', '').strip()
    update_fields = ['list_name']
    lst.list_name = new_name
    if new_status in ('active', 'inactive'):
        lst.status = new_status
        update_fields.append('status')
    lst.save(update_fields=update_fields)
    return JsonResponse({'status': 'ok', 'name': lst.list_name, 'list_status': lst.status})


def list_duplicate(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)
    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)
    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found.'}, status=404)
    base    = f'Copy of {lst.list_name}'[:255]
    name    = base
    counter = 2
    while CampaignList.objects.filter(user_id=user_id, list_name__iexact=name, deleted_at__isnull=True).exists():
        name = f'{base} ({counter})'[:255]
        counter += 1
    copy = CampaignList.objects.create(user_id=user_id, list_name=name, tags=lst.tags or '')
    return JsonResponse({'status': 'ok', 'id': copy.id, 'name': copy.list_name})


def list_download(request, list_id):
    if not request.session.get('logged_in'):
        return redirect(reverse('login'))
    from Email_validate_app.models import CampaignList, CampaignEmail
    from django.http import HttpResponse
    user_id = get_user_id(request)
    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return HttpResponse('Not found', status=404)
    rows = CampaignEmail.objects.filter(list_id=list_id, user_id=user_id, deleted_at__isnull=True)\
        .values('email', 'first_name', 'last_name', 'subscribed', 'created_at').order_by('email')
    safe_name = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in lst.list_name)[:50]
    response  = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="list_{safe_name}_{list_id}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Email', 'First Name', 'Last Name', 'Status', 'Date Added'])
    for r in rows:
        writer.writerow([
            r['email'], r['first_name'] or '', r['last_name'] or '',
            r['subscribed'], r['created_at'].strftime('%Y-%m-%d') if r['created_at'] else '',
        ])
    return response


def list_toggle_status(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)
    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)
    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)
    lst.status = 'inactive' if lst.status == 'active' else 'active'
    lst.save(update_fields=['status'])
    return JsonResponse({'status': 'ok', 'new_status': lst.status})


@_require_POST
def delete_campaign_list(request):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)

    import json as _json
    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)
    list_id = _json.loads(request.body).get('list_id')

    try:
        obj = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
        obj.deleted_at = _now()
        obj.save(update_fields=['deleted_at'])
        return JsonResponse({'status': 'ok'})
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)


def campaign_list_check(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error'}, status=403)

    from Email_validate_app.models import CampaignList, CampaignEmail
    user_id = get_user_id(request)

    try:
        CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)

    has_contacts = CampaignEmail.objects.filter(list_id=list_id, deleted_at__isnull=True).exists()
    return JsonResponse({'has_contacts': has_contacts})


def campaign_list_contacts(request, list_id):
    if not request.session.get('logged_in'):
        messages.warning(request, "You need to login first.")
        return redirect(reverse('login'))

    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)

    lst = get_object_or_404(CampaignList, id=list_id, user_id=user_id, deleted_at__isnull=True)
    lst.tags_list = [t.strip() for t in lst.tags.split(',') if t.strip()] if lst.tags else []

    return render(request, 'i_Campaign_Contacts.html', {
        'lst':                    lst,
        'pf_statuses':            CONTACT_STATUSES,
        'pf_show_status':         True,
        'pf_search_placeholder':  'Search by name or email…',
    })


def campaign_contacts_page(request, list_id):
    """AJAX endpoint — filtered, paginated contacts for a campaign list."""
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error'}, status=403)

    import traceback as _tb
    from Email_validate_app.models import CampaignList, CampaignEmail
    user_id = get_user_id(request)

    try:
        CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    try:
        f = extract_filter_params(request)
        per_page = min(
            int(request.GET.get('per_page', getattr(settings, 'FILTER_DEFAULT_PAGE_SIZE', 25))),
            getattr(settings, 'FILTER_MAX_PAGE_SIZE', 100),
        )

        qs_base = CampaignEmail.objects.filter(list_id=list_id, deleted_at__isnull=True)
        grand_total = timed_count(qs_base, 'campaign_contacts_total', user_id)

        qs = apply_search(qs_base, f['search'], 'email', 'first_name', 'last_name')
        qs = apply_status(qs, f['status'], field='subscribed')
        qs = apply_date_range(qs, f['date_from'], f['date_to'])
        qs = qs.only('id', 'first_name', 'last_name', 'email', 'subscribed', 'created_at').order_by('-created_at')

        paginator = Paginator(qs, per_page)
        page_obj  = paginator.get_page(request.GET.get('page', 1))

        start = page_obj.start_index()
        rows = []
        for i, c in enumerate(page_obj):
            rows.append({
                'id':         c.id,
                'row_num':    start + i,
                'name':       c.display_name,
                'email':      c.email,
                'subscribed': c.subscribed,
                'date':       c.created_at.strftime('%b %d, %Y'),
            })

        stats = {
            'total':            grand_total,
            'subscribed':       timed_count(qs_base.filter(subscribed='subscribed'),       'cc_sub',   user_id),
            'never_subscribed': timed_count(qs_base.filter(subscribed='never_subscribed'), 'cc_never', user_id),
            'unsubscribed':     timed_count(qs_base.filter(subscribed='unsubscribed'),     'cc_unsub', user_id),
        }

        return JsonResponse({
            'status':       'ok',
            'contacts':     rows,
            'total':        paginator.count,
            'grand_total':  grand_total,
            'stats':        stats,
            'page':         page_obj.number,
            'num_pages':    paginator.num_pages,
            'has_previous': page_obj.has_previous(),
            'has_next':     page_obj.has_next(),
            'prev_page':    page_obj.number - 1 if page_obj.has_previous() else None,
            'next_page':    page_obj.number + 1 if page_obj.has_next()     else None,
            'start_index':  start,
            'end_index':    page_obj.end_index(),
        })
    except Exception:
        return JsonResponse({'status': 'error', 'message': _tb.format_exc()}, status=500)


@_require_POST
def add_campaign_contact(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)

    from Email_validate_app.models import CampaignList, CampaignEmail
    import re, json as _json
    user_id = get_user_id(request)

    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    _body      = _json.loads(request.body)
    email      = _body.get('email', '').strip().lower()
    if not email or not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return JsonResponse({'status': 'error', 'message': 'Valid email address is required.'}, status=400)

    if CampaignEmail.objects.filter(list_id=list_id, email=email, deleted_at__isnull=True).exists():
        return JsonResponse({'status': 'error', 'message': f'{email} is already in this list.'}, status=400)

    from Email_validate_app.models import CampaignEvent

    first_name  = _body.get('first_name', '').strip()
    last_name   = _body.get('last_name', '').strip()
    phone       = _body.get('phone', '').strip()

    # Determine what the user intends to set
    uploaded_status = 'subscribed' if _body.get('consent') == '1' else 'never_subscribed'

    # Hard unsubscribe (CloudWatch / link click) always wins
    if CampaignEvent.objects.filter(user_id=user_id, email=email, event_type='unsubscribe').exists():
        existing_status = 'unsubscribed'
    else:
        # Highest-priority status this contact has in any other list
        _PRIORITY = {'unsubscribed': 3, 'never_subscribed': 2, 'subscribed': 1}
        existing_status = None
        for s in CampaignEmail.objects.filter(
            user_id=user_id, email=email, deleted_at__isnull=True
        ).exclude(list_id=list_id).values_list('subscribed', flat=True):
            if _PRIORITY.get(s, 0) > _PRIORITY.get(existing_status or '', 0):
                existing_status = s

    # Resolution matrix — lower consent always wins
    _MATRIX = {
        ('subscribed',       'subscribed'):       'subscribed',
        ('subscribed',       'unsubscribed'):     'unsubscribed',
        ('subscribed',       'never_subscribed'): 'never_subscribed',
        ('never_subscribed', 'subscribed'):       'never_subscribed',
        ('never_subscribed', 'unsubscribed'):     'unsubscribed',
        ('never_subscribed', 'never_subscribed'): 'never_subscribed',
    }
    if existing_status is None:
        contact_status = uploaded_status
    else:
        contact_status = _MATRIX.get((uploaded_status, existing_status), uploaded_status)

    CampaignEmail.objects.create(
        user_id=user_id,
        list=lst,
        first_name=first_name,
        last_name=last_name,
        email=email,
        subscribed=contact_status,
        extra_data={'phone': phone} if phone else {},
    )
    _sync_campaign_list_counts(list_id)

    return JsonResponse({'status': 'ok', 'redirect': f'/Email_Campaigns/list/{list_id}/contacts/'})


@_require_POST
def upload_campaign_contacts(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)

    from Email_validate_app.models import CampaignList, CampaignEmail, CampaignEvent
    import csv, io, re
    user_id = get_user_id(request)

    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    csv_file = request.FILES.get('csv_file')
    if not csv_file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)
    if not csv_file.name.endswith('.csv'):
        return JsonResponse({'status': 'error', 'message': 'Only CSV files are accepted.'}, status=400)

    try:
        text = csv_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        headers = [h.strip().lower() for h in (reader.fieldnames or [])]
        if 'email' not in headers:
            return JsonResponse({'status': 'error', 'message': 'CSV must have an "email" column.'}, status=400)

        email_re = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')

        # Emails already in this list — skip duplicates
        existing_emails = set(
            CampaignEmail.objects.filter(list_id=list_id, deleted_at__isnull=True)
            .values_list('email', flat=True)
        )

        # Hard unsubscribes from CloudWatch/link clicks — always wins, can never be re-subscribed
        hard_unsub = set(
            CampaignEvent.objects.filter(user_id=user_id, event_type='unsubscribe')
            .values_list('email', flat=True)
        )

        # Highest-priority status this contact already has in any other list
        _STATUS_PRIORITY = {'unsubscribed': 3, 'never_subscribed': 2, 'subscribed': 1}
        other_statuses = {}
        for ce in CampaignEmail.objects.filter(
            user_id=user_id, deleted_at__isnull=True
        ).exclude(list_id=list_id).values('email', 'subscribed'):
            e, s = ce['email'], ce['subscribed']
            if _STATUS_PRIORITY.get(s, 0) > _STATUS_PRIORITY.get(other_statuses.get(e, ''), 0):
                other_statuses[e] = s

        def get_existing(email):
            if email in hard_unsub:
                return 'unsubscribed'
            return other_statuses.get(email)

        # Status resolution matrix — lower consent always wins
        _MATRIX = {
            ('subscribed',       'subscribed'):       'subscribed',
            ('subscribed',       'unsubscribed'):     'unsubscribed',
            ('subscribed',       'never_subscribed'): 'never_subscribed',
            ('never_subscribed', 'subscribed'):       'never_subscribed',
            ('never_subscribed', 'unsubscribed'):     'unsubscribed',
            ('never_subscribed', 'never_subscribed'): 'never_subscribed',
            ('unsubscribed',     'subscribed'):       'never_subscribed',
            ('unsubscribed',     'unsubscribed'):     'unsubscribed',
            ('unsubscribed',     'never_subscribed'): 'unsubscribed',
        }

        def _parse_status(val):
            v = str(val).strip().lower().replace(' ', '_').replace('-', '_')
            if v in ('subscribed', 'subscribe', 'yes', 'true', '1', 'opt_in', 'optin'):
                return 'subscribed'
            if v in ('unsubscribed', 'unsubscribe', 'opt_out', 'optout'):
                return 'unsubscribed'
            if v in ('never_subscribed', 'never', 'no', 'false', '0'):
                return 'never_subscribed'
            return None

        def resolve_status(uploaded, existing):
            if existing is None:
                # Brand-new contact: unsubscribed in file = they never opted in
                return 'never_subscribed' if uploaded == 'unsubscribed' else uploaded
            return _MATRIX.get((uploaded, existing), uploaded)

        has_status_col = 'status' in headers
        to_create = []
        seen = set()
        skipped = 0

        for row in reader:
            row = {k.strip().lower(): v.strip() for k, v in row.items()}
            email = row.get('email', '').strip().lower()
            if not email or not email_re.match(email) or email in seen:
                continue
            seen.add(email)

            if email in existing_emails:
                skipped += 1
                continue

            uploaded_status = _parse_status(row.get('status', '')) if has_status_col else None
            if uploaded_status is None:
                uploaded_status = 'subscribed'

            final_status = resolve_status(uploaded_status, get_existing(email))

            phone = row.get('phone', '')
            to_create.append(CampaignEmail(
                user_id=user_id,
                list=lst,
                first_name=row.get('first_name', ''),
                last_name=row.get('last_name', ''),
                email=email,
                subscribed=final_status,
                extra_data={'phone': phone} if phone else {},
            ))

        if not to_create and skipped == 0:
            return JsonResponse({'status': 'error', 'message': 'No valid email addresses found in the file.'}, status=400)
        if not to_create:
            return JsonResponse({'status': 'error', 'message': f'All {skipped} email(s) already exist in this list.'}, status=400)

        CampaignEmail.objects.bulk_create(to_create, ignore_conflicts=True)
        _sync_campaign_list_counts(list_id)

        msg = f'{len(to_create)} contact(s) imported.'
        if skipped:
            msg += f' {skipped} duplicate(s) skipped.'
        return JsonResponse({'status': 'ok', 'imported': len(to_create), 'skipped': skipped, 'message': msg})

    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not parse file: {exc}'}, status=400)


# Upload wizard: Step 1 — parse file, return columns
@_require_POST
def parse_upload_file(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)

    from Email_validate_app.models import CampaignList
    user_id = get_user_id(request)
    try:
        CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    fname = uploaded.name.lower()
    _NO_HEADER_MSG = (
        'No header row detected. The first row must contain column names '
        '(e.g. "email", "first_name"). Please add a header row and try again.'
    )
    try:
        if fname.endswith('.csv') or fname.endswith('.txt'):
            import csv, io, re as _re
            text = uploaded.read().decode('utf-8-sig')
            if not text.strip():
                return JsonResponse({'status': 'error', 'message': 'The file is empty.'}, status=400)
            try:
                dialect = csv.Sniffer().sniff(text[:2048], delimiters=',\t;|')
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            columns = [c.strip() for c in (reader.fieldnames or []) if c and c.strip()]
            # Block only when the first row clearly contains data, not headers
            _email_re = _re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
            data_like = sum(
                1 for c in columns
                if _email_re.match(c) or c.lstrip('-').replace('.', '', 1).isdigit()
            )
            if columns and data_like > 0 and data_like >= len(columns) / 2:
                return JsonResponse({'status': 'error', 'message': _NO_HEADER_MSG}, status=400)
        elif fname.endswith('.xlsx'):
            import openpyxl, re as _re
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(min_row=1, max_row=2, values_only=True)
            header_row = next(rows_iter, [])
            second_row = next(rows_iter, [])
            wb.close()
            header_cells = [str(c).strip() for c in header_row if c is not None and str(c).strip()]
            if not header_cells:
                return JsonResponse({'status': 'error', 'message': 'No columns detected. Check the file format.'}, status=400)
            # If any header cell looks like an email or a bare number, first row is likely data
            _email_re = _re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
            data_like = sum(
                1 for c in header_cells
                if _email_re.match(c) or c.lstrip('-').replace('.', '', 1).isdigit()
            )
            if data_like > 0 and data_like >= len(header_cells) / 2:
                return JsonResponse({'status': 'error', 'message': _NO_HEADER_MSG}, status=400)
            columns = header_cells
        else:
            return JsonResponse({'status': 'error', 'message': 'Only CSV, XLSX, and TXT files are accepted.'}, status=400)

        if not columns:
            return JsonResponse({'status': 'error', 'message': 'No columns detected. Check the file format.'}, status=400)

        return JsonResponse({'status': 'ok', 'columns': columns})
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {exc}'}, status=400)


# Upload wizard: Step 2 — import with mapping
@_require_POST
def import_contacts(request, list_id):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)

    from Email_validate_app.models import CampaignList, CampaignEmail
    import csv, io, json as _json, re
    user_id = get_user_id(request)

    try:
        lst = CampaignList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except CampaignList.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    uploaded = request.FILES.get('file')
    try:
        mapping = _json.loads(request.POST.get('mapping', '{}'))  # {file_col: system_field}
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Invalid mapping data.'}, status=400)

    if 'email' not in mapping.values():
        return JsonResponse({'status': 'error', 'message': 'Email field must be mapped.'}, status=400)
    if 'status' not in mapping.values():
        return JsonResponse({'status': 'error', 'message': 'Status field must be mapped.'}, status=400)
    if 'first_name' not in mapping.values() and 'last_name' not in mapping.values():
        return JsonResponse({'status': 'error', 'message': 'At least one name field (first_name or last_name) must be mapped.'}, status=400)

    email_col      = next(c for c, s in mapping.items() if s == 'email')
    status_col     = next(c for c, s in mapping.items() if s == 'status')
    first_name_col = next((c for c, s in mapping.items() if s == 'first_name'), None)
    last_name_col  = next((c for c, s in mapping.items() if s == 'last_name'), None)
    mapped_cols    = set(mapping.keys())

    def _normalize_status(val):
        v = str(val).strip().lower().replace(' ', '_').replace('-', '_')
        if v in ('subscribed', 'subscribe', 'yes', 'true', '1', 'opt_in', 'optin'):
            return 'subscribed'
        if v in ('unsubscribed', 'unsubscribe', 'opt_out', 'optout'):
            return 'unsubscribed'
        if v in ('never_subscribed', 'never', 'no', 'false', '0'):
            return 'never_subscribed'
        return None

    # Parse file into list of dicts
    fname = uploaded.name.lower()
    try:
        rows = []
        if fname.endswith('.csv') or fname.endswith('.txt'):
            text = uploaded.read().decode('utf-8-sig')
            try:
                dialect = csv.Sniffer().sniff(text[:2048], delimiters=',\t;|')
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
        elif fname.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            header = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    header = [str(c).strip() if c is not None else f'col{j}' for j, c in enumerate(row)]
                else:
                    rows.append({header[j]: (str(v).strip() if v is not None else '') for j, v in enumerate(row)})
            wb.close()
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {exc}'}, status=400)

    # Mapping validation against a sample of rows
    email_re = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
    sample = [r for r in rows if (r.get(email_col) or '').strip()][:10]
    if sample:
        # Email column: >= 60 % of sampled values must look like email addresses
        email_hits = sum(1 for r in sample if email_re.match((r.get(email_col) or '').strip().lower()))
        if email_hits < max(1, len(sample) * 0.6):
            return JsonResponse({
                'status': 'error',
                'message': (
                    f'The column mapped to Email ("{email_col}") does not appear to contain '
                    f'email addresses (checked {len(sample)} rows, only {email_hits} looked valid). '
                    f'Please check your column mapping.'
                ),
            }, status=400)

        # First Name / Last Name: < 30 % of sampled values should look like email addresses
        for field_label, col_name in [('First Name', first_name_col), ('Last Name', last_name_col)]:
            if not col_name:
                continue
            email_like = sum(1 for r in sample if email_re.match((r.get(col_name) or '').strip().lower()))
            if email_like >= len(sample) * 0.3:
                return JsonResponse({
                    'status': 'error',
                    'message': (
                        f'The column mapped to {field_label} ("{col_name}") appears to contain '
                        f'email addresses rather than names. Please check your column mapping.'
                    ),
                }, status=400)

    existing_emails = set(
        CampaignEmail.objects.filter(list_id=list_id, deleted_at__isnull=True)
        .values_list('email', flat=True)
    )

    from Email_validate_app.models import CampaignEvent

    # Build existing-status map across all other lists
    hard_unsub = set(
        CampaignEvent.objects.filter(user_id=user_id, event_type='unsubscribe')
        .values_list('email', flat=True)
    )
    _STATUS_PRIORITY = {'unsubscribed': 3, 'never_subscribed': 2, 'subscribed': 1}
    other_statuses = {}
    for _ce in CampaignEmail.objects.filter(
        user_id=user_id, deleted_at__isnull=True
    ).exclude(list_id=list_id).values('email', 'subscribed'):
        _e, _s = _ce['email'], _ce['subscribed']
        if _STATUS_PRIORITY.get(_s, 0) > _STATUS_PRIORITY.get(other_statuses.get(_e, ''), 0):
            other_statuses[_e] = _s

    def get_existing(email):
        if email in hard_unsub:
            return 'unsubscribed'
        return other_statuses.get(email)

    # Status resolution matrix — lower consent always wins
    _MATRIX = {
        ('subscribed',       'subscribed'):       'subscribed',
        ('subscribed',       'unsubscribed'):     'unsubscribed',
        ('subscribed',       'never_subscribed'): 'never_subscribed',
        ('never_subscribed', 'subscribed'):       'never_subscribed',
        ('never_subscribed', 'unsubscribed'):     'unsubscribed',
        ('never_subscribed', 'never_subscribed'): 'never_subscribed',
        ('unsubscribed',     'subscribed'):       'never_subscribed',
        ('unsubscribed',     'unsubscribed'):     'unsubscribed',
        ('unsubscribed',     'never_subscribed'): 'unsubscribed',
    }

    def resolve_status(uploaded, existing):
        if existing is None:
            # Brand-new contact: unsubscribed in file means they never opted in
            return 'never_subscribed' if uploaded == 'unsubscribed' else uploaded
        return _MATRIX.get((uploaded, existing), uploaded)

    # Pre-flight conflict detection
    if not request.POST.get('confirmed'):
        conflicts = []
        for i, row in enumerate(rows):
            email = (row.get(email_col) or '').strip().lower()
            if not email or not email_re.match(email) or email in existing_emails:
                continue
            uploaded_status = _normalize_status(row.get(status_col) or '')
            if uploaded_status is None:
                continue
            existing_status = get_existing(email)
            if existing_status is None:
                continue  # brand-new contact: status rule is silent, no conflict to review
            final = resolve_status(uploaded_status, existing_status)
            if final != uploaded_status:
                conflicts.append({
                    'row':      i + 2,
                    'email':    email,
                    'uploaded': uploaded_status.replace('_', ' '),
                    'existing': (existing_status or '').replace('_', ' '),
                    'final':    final.replace('_', ' '),
                })
        if conflicts:
            return JsonResponse({
                'status':          'conflicts',
                'conflicts':       conflicts[:100],
                'total_conflicts': len(conflicts),
            })

    # Import
    to_create, seen = [], set()
    total = len(rows)
    skipped_dup = skipped_invalid = 0
    imported_sub = imported_neversub = imported_unsub = 0
    skipped_rows = []

    for i, row in enumerate(rows):
        raw_email      = (row.get(email_col)  or '').strip().lower()
        raw_status_val = (row.get(status_col) or '').strip()

        if not raw_email:
            skipped_rows.append({'row': i + 2, 'email': '(empty)', 'uploaded_status': raw_status_val, 'reason': 'Missing email'})
            skipped_invalid += 1
            continue
        if not email_re.match(raw_email):
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status_val, 'reason': 'Invalid email format'})
            skipped_invalid += 1
            continue
        if raw_email in seen or raw_email in existing_emails:
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status_val, 'reason': 'Already exists in this list'})
            skipped_dup += 1
            continue

        uploaded_status = _normalize_status(raw_status_val)
        if uploaded_status is None:
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status_val, 'reason': f'Invalid status value "{raw_status_val}"'})
            skipped_invalid += 1
            continue

        final_status = resolve_status(uploaded_status, get_existing(raw_email))
        seen.add(raw_email)
        extra = {col: (row.get(col) or '').strip()
                 for col in row if col not in mapped_cols and (row.get(col) or '').strip()}
        to_create.append(CampaignEmail(
            user_id=user_id,
            list=lst,
            first_name=(row.get(first_name_col) or '').strip() if first_name_col else '',
            last_name=(row.get(last_name_col)  or '').strip()  if last_name_col  else '',
            email=raw_email,
            subscribed=final_status,
            extra_data=extra,
        ))
        if final_status == 'subscribed':
            imported_sub += 1
        elif final_status == 'unsubscribed':
            imported_unsub += 1
        else:
            imported_neversub += 1

    CampaignEmail.objects.bulk_create(to_create, ignore_conflicts=True)
    _sync_campaign_list_counts(list_id)

    imported = len(to_create)
    return JsonResponse({
        'status': 'ok',
        'summary': {
            'total':    total,
            'imported': imported,
            'by_status': {
                'subscribed':       imported_sub,
                'unsubscribed':     imported_unsub,
                'never_subscribed': imported_neversub,
            },
            'skipped':      len(skipped_rows),
            'skipped_rows': skipped_rows,
        }
    })


# ── All Contacts ──────────────────────────────────────────────────────────────

def all_contacts(request):
    """Render the All Contacts page (deduped across lists)."""
    if not request.session.get('logged_in'):
        messages.warning(request, "You need to login first.")
        return redirect(reverse('login'))
    return render(request, 'i_All_Contacts.html', {
        'pf_statuses': CONTACT_STATUSES,
    })


def _normalize_sub(val):
    """Normalize legacy '0'/'1' bool-string values to canonical status choices."""
    if val in ('subscribed', '1'):
        return 'subscribed'
    if val == 'unsubscribed':
        return 'unsubscribed'
    return 'never_subscribed'  # '0', 'never_subscribed', or any unknown value


# DB values that map to each canonical status (legacy '0'/'1' + new strings)
_STATUS_DB_VALUES = {
    'subscribed':       ['subscribed', '1'],
    'unsubscribed':     ['unsubscribed'],
    'never_subscribed': ['never_subscribed', '0'],
}


def all_contacts_page(request):
    """AJAX endpoint — paginated, filtered list of unique contacts."""
    if not request.session.get('logged_in'):
        return JsonResponse({'error': 'Not authenticated'}, status=403)

    from Email_validate_app.models import CampaignEmail
    from django.db.models import Count as _Count
    user_id = get_user_id(request)

    f        = extract_filter_params(request)
    search   = f['search']
    status   = f['status']
    d_from   = f['date_from']
    d_to     = f['date_to']
    page     = max(1, int(request.GET.get('page', 1)))
    per_page = min(
        int(request.GET.get('per_page', getattr(settings, 'FILTER_DEFAULT_PAGE_SIZE', 25))),
        getattr(settings, 'FILTER_MAX_PAGE_SIZE', 100),
    )

    # Step 1: build the global canonical ID list (one record per unique email,
    # the earliest id regardless of status). This is the single source of truth
    # for both the stat counts and the filter — keeping them consistent.
    all_canonical_ids = list(
        CampaignEmail.objects
        .filter(user_id=user_id, deleted_at__isnull=True)
        .values('email')
        .annotate(first_id=Min('id'))
        .values_list('first_id', flat=True)
    )
    grand_total = len(all_canonical_ids)

    # Step 2: start from canonical records and apply search / status / date filters
    base_qs = CampaignEmail.objects.filter(id__in=all_canonical_ids)

    if search:
        base_qs = base_qs.filter(
            Q(email__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search)
        )
    if status and status != 'all':
        # Match both legacy ('0'/'1') and canonical string values
        base_qs = base_qs.filter(subscribed__in=_STATUS_DB_VALUES.get(status, [status]))
    if d_from:
        base_qs = base_qs.filter(created_at__date__gte=d_from)
    if d_to:
        base_qs = base_qs.filter(created_at__date__lte=d_to)

    base_qs = base_qs.order_by('-created_at')
    total  = base_qs.count()
    offset = (page - 1) * per_page
    rows   = list(base_qs.values('id', 'first_name', 'last_name', 'email', 'subscribed', 'created_at')[offset:offset + per_page])

    # Step 3: build email → [list_names] mapping for the current page emails
    page_emails = [r['email'] for r in rows]
    memberships = (
        CampaignEmail.objects
        .filter(user_id=user_id, email__in=page_emails, deleted_at__isnull=True)
        .select_related('list')
        .values('email', 'list__list_name')
    )
    email_lists: dict = {}
    for m in memberships:
        lst_name = m['list__list_name'] or ''
        if lst_name:
            email_lists.setdefault(m['email'], [])
            if lst_name not in email_lists[m['email']]:
                email_lists[m['email']].append(lst_name)

    # Step 4: format rows — normalize legacy status values
    for r in rows:
        r['name']      = f"{r['first_name'] or ''} {r['last_name'] or ''}".strip() or '—'
        r['subscribed'] = _normalize_sub(r['subscribed'])
        r['lists']      = email_lists.get(r['email'], [])
        r['created_at'] = r['created_at'].strftime('%d %b %Y') if r['created_at'] else '—'
        del r['first_name'], r['last_name']

    # Step 5: stat counts — always from all canonical records (no search/status filter)
    # Using the same all_canonical_ids ensures stats match what the filter would return
    raw_counts = {
        row['subscribed']: row['cnt']
        for row in CampaignEmail.objects
            .filter(id__in=all_canonical_ids)
            .values('subscribed')
            .annotate(cnt=_Count('id'))
    }
    normalized = {'subscribed': 0, 'unsubscribed': 0, 'never_subscribed': 0}
    for val, cnt in raw_counts.items():
        normalized[_normalize_sub(val)] += cnt
    total_all = sum(normalized.values())

    return JsonResponse({
        'contacts':    rows,
        'total':       total,
        'grand_total': grand_total,
        'page':        page,
        'per_page':    per_page,
        'pages':       max(1, (total + per_page - 1) // per_page),
        'stats': {
            'total':            total_all,
            'subscribed':       normalized['subscribed'],
            'unsubscribed':     normalized['unsubscribed'],
            'never_subscribed': normalized['never_subscribed'],
        },
    })


def contact_detail(request, contact_id):
    """Return full detail JSON for a single contact."""
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)

    from Email_validate_app.models import CampaignEmail, CampaignEvent
    user_id = get_user_id(request)

    try:
        contact = CampaignEmail.objects.select_related('list').get(
            id=contact_id, user_id=user_id, deleted_at__isnull=True,
        )
    except CampaignEmail.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Contact not found.'}, status=404)

    extra = contact.extra_data or {}
    hard_unsub = CampaignEvent.objects.filter(
        user_id=user_id, email=contact.email, event_type='unsubscribe',
    ).exists()

    return JsonResponse({
        'status': 'ok',
        'contact': {
            'id':         contact.id,
            'first_name': contact.first_name,
            'last_name':  contact.last_name,
            'email':      contact.email,
            'subscribed': contact.subscribed,
            'hard_unsub': hard_unsub,
            'phone':      extra.get('phone', ''),
            'company':    extra.get('company', ''),
            'extra_data': {k: v for k, v in extra.items() if k not in ('phone', 'company')},
            'list_id':    contact.list_id,
            'list_name':  contact.list.list_name,
            'created_at': contact.created_at.strftime('%d %b %Y, %H:%M'),
            'updated_at': contact.updated_at.strftime('%d %b %Y, %H:%M'),
        },
    })
