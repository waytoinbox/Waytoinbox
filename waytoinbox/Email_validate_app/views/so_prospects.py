import csv
import io
import json
import re
from datetime import datetime

from django.core.paginator import Paginator
from django.db.models import Count
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.timezone import now

from Email_validate_app.utils import get_user_id


def _auth(request):
    if not request.session.get('logged_in'):
        return redirect(reverse('login'))


def so_prospects(request):
    r = _auth(request)
    if r:
        return r
    from Email_validate_app.models import SOProspect
    user_id  = get_user_id(request)
    base_qs  = SOProspect.objects.filter(user_id=user_id, deleted_at__isnull=True)
    qs       = base_qs.prefetch_related('prospect_lists__so_list')

    search    = request.GET.get('q', '').strip()
    status    = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(email__icontains=search) | Q(first_name__icontains=search) |
            Q(last_name__icontains=search) | Q(company__icontains=search)
        )
    if status:
        qs = qs.filter(status=status)

    def _parse_date(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None

    parsed_from = _parse_date(date_from)
    parsed_to   = _parse_date(date_to)
    if parsed_from:
        qs = qs.filter(created_at__date__gte=parsed_from)
    if parsed_to:
        qs = qs.filter(created_at__date__lte=parsed_to)

    try:
        page_size = int(request.GET.get('page_size') or 5)
    except (TypeError, ValueError):
        page_size = 5
    page_size = max(1, min(100, page_size))

    paginator  = Paginator(qs, page_size)
    page_obj   = paginator.get_page(request.GET.get('page', 1))
    total      = qs.count()

    # Stat cards always reflect every prospect (ignoring search/status/date
    # filters), matching All Contacts' behavior — the counts are a stable
    # summary of the whole list, not a reflection of the current filter.
    raw_counts = {
        row['status']: row['cnt']
        for row in base_qs.values('status').annotate(cnt=Count('id'))
    }
    stats = {
        'total':            sum(raw_counts.values()),
        'subscribed':       raw_counts.get('subscribed', 0),
        'unsubscribed':     raw_counts.get('unsubscribed', 0),
        'never_subscribed': raw_counts.get('never_subscribed', 0),
    }

    return render(request, 'i_SO_Prospects.html', {
        'page_obj': page_obj,
        'total': total,
        'search': search,
        'status_filter': status,
        'date_from': date_from,
        'date_to': date_to,
        'page_size': page_size,
        'stats': stats,
    })


def so_prospects_action(request):
    r = _auth(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    from Email_validate_app.models import SOProspect
    data    = json.loads(request.body)
    action  = data.get('action')
    user_id = get_user_id(request)

    if action == 'add':
        email = (data.get('email') or '').strip().lower()
        if not email or '@' not in email:
            return JsonResponse({'status': 'error', 'message': 'Valid email required.'})
        p, created = SOProspect.objects.update_or_create(
            user_id=user_id, email=email,
            defaults={
                'first_name': (data.get('first_name') or '').strip(),
                'last_name':  (data.get('last_name')  or '').strip(),
                'company':    (data.get('company')    or '').strip(),
                'phone':      (data.get('phone')      or '').strip(),
                'deleted_at': None,
            },
        )
        return JsonResponse({'status': 'ok', 'id': p.id, 'created': created})

    if action == 'update':
        pid = data.get('id')
        try:
            p = SOProspect.objects.get(id=pid, user_id=user_id, deleted_at__isnull=True)
        except SOProspect.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Not found.'})
        for field in ('first_name', 'last_name', 'company', 'phone'):
            if field in data:
                setattr(p, field, (data[field] or '').strip())
        p.save()
        return JsonResponse({'status': 'ok'})

    if action == 'delete':
        pid = data.get('id')
        SOProspect.objects.filter(id=pid, user_id=user_id).update(deleted_at=now())
        return JsonResponse({'status': 'ok'})

    if action == 'bulk_delete':
        ids = data.get('ids', [])
        SOProspect.objects.filter(id__in=ids, user_id=user_id).update(deleted_at=now())
        return JsonResponse({'status': 'ok'})

    if action == 'search':
        from django.db.models import Q
        q     = (data.get('q') or '').strip()
        limit = int(data.get('limit') or 30)
        qs    = SOProspect.objects.filter(user_id=user_id, deleted_at__isnull=True)
        if q:
            qs = qs.filter(
                Q(email__icontains=q) | Q(first_name__icontains=q) |
                Q(last_name__icontains=q) | Q(company__icontains=q)
            )
        results = [
            {'id': p.id, 'email': p.email,
             'name': f'{p.first_name} {p.last_name}'.strip()}
            for p in qs[:limit]
        ]
        return JsonResponse({'status': 'ok', 'results': results})

    return JsonResponse({'status': 'error', 'message': 'Unknown action.'})


_NO_HEADER_MSG = (
    'No header row detected. The first row must contain column names '
    '(e.g. "email", "first_name"). Please add a header row and try again.'
)
_EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')


def _read_upload_rows(uploaded):
    """Parse an uploaded CSV/TXT/XLSX into a list of dicts keyed by column
    name. Same shape as services/so_lists.py::_read_rows (not imported
    directly — file-parsing boilerplate is duplicated per-view elsewhere in
    this codebase too, e.g. contacts.py vs so_lists.py, rather than shared
    across view modules)."""
    fname = uploaded.name.lower()
    rows = []
    if fname.endswith('.csv') or fname.endswith('.txt'):
        text = uploaded.read().decode('utf-8-sig', errors='replace')
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=',\t;|')
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
    elif fname.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c).strip() if c is not None else f'col{j}' for j, c in enumerate(row)]
            else:
                rows.append({
                    header[j]: (str(v).strip() if v is not None else '')
                    for j, v in enumerate(row) if j < len(header)
                })
        wb.close()
    return rows


def so_prospects_parse_file(request):
    """Upload wizard step 1 — read the uploaded file and return its column names."""
    r = _auth(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    fname = uploaded.name.lower()
    try:
        if fname.endswith('.csv') or fname.endswith('.txt'):
            text = uploaded.read().decode('utf-8-sig', errors='replace')
            if not text.strip():
                return JsonResponse({'status': 'error', 'message': 'The file is empty.'}, status=400)
            try:
                dialect = csv.Sniffer().sniff(text[:2048], delimiters=',\t;|')
            except csv.Error:
                dialect = csv.excel
            reader  = csv.DictReader(io.StringIO(text), dialect=dialect)
            columns = [c.strip() for c in (reader.fieldnames or []) if c and c.strip()]
        elif fname.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            wb.close()
            columns = [str(c).strip() for c in header_row if c is not None and str(c).strip()]
        else:
            return JsonResponse({'status': 'error', 'message': 'Only CSV, XLSX, and TXT files are accepted.'}, status=400)

        if not columns:
            return JsonResponse({'status': 'error', 'message': 'No columns detected. Check the file format.'}, status=400)

        data_like = sum(
            1 for c in columns
            if _EMAIL_RE.match(c) or c.lstrip('-').replace('.', '', 1).isdigit()
        )
        if data_like > 0 and data_like >= len(columns) / 2:
            return JsonResponse({'status': 'error', 'message': _NO_HEADER_MSG}, status=400)

        return JsonResponse({'status': 'ok', 'columns': columns})
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {exc}'}, status=400)


def so_prospects_import(request):
    """Upload wizard step 2 — import the file using the column mapping.

    Mirrors views/so_lists.py::so_list_import_prospects's wizard shape and
    validation exactly (mapping validation, email/name sanity checks,
    conflict pre-flight, per-row skip reasons) — _normalize_status and
    _resolve_status are imported from there directly rather than
    re-implemented, since the consent-protection matrix (an unsubscribed
    prospect can never be silently re-subscribed by a re-upload) must never
    drift between the two import paths. The one genuine behavioral
    difference: this page has no list context, so an already-existing
    prospect is updated in place rather than skipped as "already in list"
    (matching this endpoint's pre-existing upsert-by-email semantics)."""
    r = _auth(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    from Email_validate_app.models import SOProspect, SOEvent
    from Email_validate_app.views.so_lists import _normalize_status, _resolve_status

    user_id = get_user_id(request)
    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    try:
        mapping = json.loads(request.POST.get('mapping', '{}'))  # {file_col: system_field}
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Invalid mapping data.'}, status=400)

    if 'email' not in mapping.values():
        return JsonResponse({'status': 'error', 'message': 'Email field must be mapped.'}, status=400)
    if 'first_name' not in mapping.values() and 'last_name' not in mapping.values():
        return JsonResponse({
            'status': 'error',
            'message': 'At least one name field (first_name or last_name) must be mapped.',
        }, status=400)

    email_col      = next(c for c, s in mapping.items() if s == 'email')
    status_col     = next((c for c, s in mapping.items() if s == 'status'),     None)
    first_name_col = next((c for c, s in mapping.items() if s == 'first_name'), None)
    last_name_col  = next((c for c, s in mapping.items() if s == 'last_name'),  None)
    company_col    = next((c for c, s in mapping.items() if s == 'company'),    None)
    phone_col      = next((c for c, s in mapping.items() if s == 'phone'),      None)
    mapped_cols    = set(mapping.keys())

    try:
        rows = _read_upload_rows(uploaded)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {exc}'}, status=400)

    # ── Sanity-check the mapping against a sample of real rows ────────────────
    sample = [r for r in rows if (r.get(email_col) or '').strip()][:10]
    if sample:
        hits = sum(1 for r in sample if _EMAIL_RE.match((r.get(email_col) or '').strip().lower()))
        if hits < max(1, len(sample) * 0.6):
            return JsonResponse({
                'status': 'error',
                'message': (
                    f'The column mapped to Email ("{email_col}") does not appear to contain email '
                    f'addresses (checked {len(sample)} rows, only {hits} looked valid). '
                    f'Please check your column mapping.'
                ),
            }, status=400)
        for label, col in [('First Name', first_name_col), ('Last Name', last_name_col)]:
            if not col:
                continue
            email_like = sum(1 for r in sample if _EMAIL_RE.match((r.get(col) or '').strip().lower()))
            if email_like >= len(sample) * 0.3:
                return JsonResponse({
                    'status': 'error',
                    'message': (
                        f'The column mapped to {label} ("{col}") appears to contain email addresses '
                        f'rather than names. Please check your column mapping.'
                    ),
                }, status=400)

    # ── Existing state ──────────────────────────────────────────────────────
    existing_prospects = {
        p['email']: p for p in SOProspect.objects.filter(user_id=user_id)
        .values('id', 'email', 'status', 'deleted_at')
    }
    # Tracked unsubscribes always win and can never be re-subscribed.
    hard_unsub = set(
        SOEvent.objects.filter(campaign__user_id=user_id, event_type='unsubscribed')
        .values_list('email', flat=True)
    )

    def existing_status(email):
        # Deliberately ignores deleted_at — an unsubscribe is a consent
        # signal that survives a soft-delete of the prospect record.
        if email in hard_unsub:
            return 'unsubscribed'
        p = existing_prospects.get(email)
        return p['status'] if p else None

    # ── Pre-flight: rows whose status will be overridden by existing state ────
    if not request.POST.get('confirmed'):
        conflicts = []
        for i, row in enumerate(rows):
            email = (row.get(email_col) or '').strip().lower()
            if not email or not _EMAIL_RE.match(email):
                continue
            uploaded_status = _normalize_status(row.get(status_col)) if status_col else 'subscribed'
            if uploaded_status is None:
                continue
            existing = existing_status(email)
            if existing is None:
                continue
            final = _resolve_status(uploaded_status, existing)
            if final != uploaded_status:
                conflicts.append({
                    'row':      i + 2,
                    'email':    email,
                    'uploaded': uploaded_status,
                    'existing': existing,
                    'final':    final,
                })
        if conflicts:
            return JsonResponse({
                'status':          'conflicts',
                'conflicts':       conflicts[:100],
                'total_conflicts': len(conflicts),
            })

    # ── Import ──────────────────────────────────────────────────────────────
    total        = len(rows)
    seen         = set()
    skipped_rows = []
    to_create    = []
    to_update    = []
    by_status    = {'subscribed': 0, 'unsubscribed': 0, 'never_subscribed': 0}

    for i, row in enumerate(rows):
        raw_email  = (row.get(email_col) or '').strip().lower()
        raw_status = (row.get(status_col) or '').strip() if status_col else ''

        if not raw_email:
            skipped_rows.append({'row': i + 2, 'email': '(empty)', 'uploaded_status': raw_status, 'reason': 'Missing email'})
            continue
        if not _EMAIL_RE.match(raw_email):
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status, 'reason': 'Invalid email format'})
            continue
        if raw_email in seen:
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status, 'reason': 'Duplicate row in file'})
            continue

        if status_col:
            uploaded_status = _normalize_status(raw_status)
            if uploaded_status is None:
                skipped_rows.append({
                    'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status,
                    'reason': f'Invalid status value "{raw_status}"',
                })
                continue
        else:
            uploaded_status = 'subscribed'

        seen.add(raw_email)
        final_status = _resolve_status(uploaded_status, existing_status(raw_email))
        by_status[final_status] = by_status.get(final_status, 0) + 1

        first = (row.get(first_name_col) or '').strip() if first_name_col else ''
        last  = (row.get(last_name_col)  or '').strip() if last_name_col  else ''
        comp  = (row.get(company_col)    or '').strip() if company_col    else ''
        phone = (row.get(phone_col)      or '').strip() if phone_col      else ''
        extra = {
            col: (row.get(col) or '').strip()
            for col in row if col not in mapped_cols and (row.get(col) or '').strip()
        }

        known = existing_prospects.get(raw_email)
        if known:
            to_update.append((known['id'], first, last, comp, phone, final_status, extra))
        else:
            to_create.append(SOProspect(
                user_id=user_id, email=raw_email, first_name=first, last_name=last,
                company=comp, phone=phone, status=final_status, extra_data=extra,
            ))

    if to_create:
        SOProspect.objects.bulk_create(to_create, ignore_conflicts=True)

    for pid, first, last, comp, phone, final_status, extra in to_update:
        p = SOProspect.objects.filter(id=pid).first()
        if not p:
            continue
        p.first_name = first or p.first_name
        p.last_name  = last  or p.last_name
        p.company    = comp  or p.company
        p.phone      = phone or p.phone
        p.status     = final_status
        p.deleted_at = None
        if extra:
            merged = dict(p.extra_data or {})
            merged.update(extra)
            p.extra_data = merged
        p.save()

    imported = len(seen)
    return JsonResponse({
        'status': 'ok',
        'summary': {
            'total':        total,
            'imported':     imported,
            'by_status':    by_status,
            'skipped':      len(skipped_rows),
            'skipped_rows': skipped_rows[:500],
        },
    })
