import csv
import io
import json
import re

from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils.timezone import now as _now
from django.views.decorators.http import require_POST as _require_POST

from Email_validate_app.utils import get_user_id
from Email_validate_app.services.filter_utils import (
    extract_filter_params, apply_search, apply_status, apply_date_range, timed_count,
)
from Email_validate_app.services.filter_status import CONTACT_STATUSES


EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')

# Highest-priority status a prospect already holds anywhere wins over the
# uploaded one. Identical to the Email Marketing rule.
_STATUS_PRIORITY = {'unsubscribed': 3, 'never_subscribed': 2, 'subscribed': 1}

# Status resolution matrix — lower consent always wins. Copied verbatim from
# the Email Marketing importer so both products behave identically.
_MATRIX = {
    ('subscribed',       'subscribed'):       'subscribed',
    ('subscribed',       'unsubscribed'):     'unsubscribed',
    ('subscribed',       'never_subscribed'): 'never_subscribed',
    ('never_subscribed', 'subscribed'):       'never_subscribed',
    ('never_subscribed', 'unsubscribed'):     'unsubscribed',
    ('never_subscribed', 'never_subscribed'): 'never_subscribed',
    ('unsubscribed',     'subscribed'):       'never_subscribed',
    ('unsubscribed',     'unsubscribed'):     'unsubscribed',
    ('unsubscribed',     'never_subscribed'): 'unsubscribed',
}


def _auth(request):
    if not request.session.get('logged_in'):
        messages.warning(request, "You need to login first.")
        return redirect(reverse('login'))


def _auth_json(request):
    if not request.session.get('logged_in'):
        return JsonResponse({'status': 'error', 'message': 'Not authenticated'}, status=403)


def _get_list(user_id, list_id):
    from Email_validate_app.models import SOList
    try:
        return SOList.objects.get(id=list_id, user_id=user_id, deleted_at__isnull=True)
    except SOList.DoesNotExist:
        return None


def _sync_so_list_counts(list_id):
    """Recompute the cached prospect counters on a list."""
    from Email_validate_app.models import SOList, SOListProspect
    qs = SOListProspect.objects.filter(so_list_id=list_id, prospect__deleted_at__isnull=True)
    SOList.objects.filter(id=list_id).update(
        total_count=qs.count(),
        subscribed_count=qs.filter(prospect__status='subscribed').count(),
        neversubscribed_count=qs.filter(prospect__status='never_subscribed').count(),
        unsubscribed_count=qs.filter(prospect__status='unsubscribed').count(),
    )


def _normalize_status(val):
    """Map a free-text status value from a file to a canonical prospect status."""
    v = str(val or '').strip().lower().replace(' ', '_').replace('-', '_')
    if v in ('subscribed', 'subscribe', 'yes', 'true', '1', 'opt_in', 'optin', 'active'):
        return 'subscribed'
    if v in ('unsubscribed', 'unsubscribe', 'opt_out', 'optout', 'bounced', 'bounce'):
        return 'unsubscribed'
    if v in ('never_subscribed', 'never', 'no', 'false', '0'):
        return 'never_subscribed'
    return None


def _resolve_status(uploaded, existing):
    """Lower consent always wins — same rule as the Email Marketing importer."""
    if existing is None:
        # Brand-new prospect: 'unsubscribed' in the file means they never opted in
        return 'never_subscribed' if uploaded == 'unsubscribed' else uploaded
    return _MATRIX.get((uploaded, existing), uploaded)


# ── Lists page ────────────────────────────────────────────────────────────────

def so_lists(request):
    """Render the Lists page; POST creates a new list."""
    r = _auth(request)
    if r:
        return r

    from Email_validate_app.models import SOList
    user_id = get_user_id(request)

    if request.method == 'POST':
        try:
            body = json.loads(request.body)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)
        name = (body.get('name') or body.get('list_name') or '').strip()
        tags = (body.get('tags') or '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'List name is required.'}, status=400)
        if len(name) > 255:
            return JsonResponse({'success': False, 'error': 'Name too long.'}, status=400)
        if SOList.objects.filter(user_id=user_id, name__iexact=name, deleted_at__isnull=True).exists():
            return JsonResponse({'success': False, 'error': f'A list named "{name}" already exists.'}, status=400)
        lst = SOList.objects.create(user_id=user_id, name=name, tags=tags)
        return JsonResponse({'success': True, 'id': lst.id, 'name': lst.name})

    lists = SOList.objects.filter(user_id=user_id, deleted_at__isnull=True).order_by('-created_at')
    for lst in lists:
        lst.tags_list = [t.strip() for t in lst.tags.split(',') if t.strip()] if lst.tags else []
    return render(request, 'i_SO_Lists.html', {
        'lists':       lists,
        'pf_statuses': [('active', 'Active'), ('inactive', 'Inactive')],
    })


def so_list_rename(request, list_id):
    """Edit a list's name, tags and/or status."""
    r = _auth_json(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    from Email_validate_app.models import SOList
    user_id = get_user_id(request)
    try:
        body = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    new_name = (body.get('name') or '').strip()
    if not new_name:
        return JsonResponse({'status': 'error', 'message': 'Name is required.'}, status=400)
    if len(new_name) > 255:
        return JsonResponse({'status': 'error', 'message': 'Name too long.'}, status=400)

    lst = _get_list(user_id, list_id)
    if not lst:
        return JsonResponse({'status': 'error', 'message': 'List not found.'}, status=404)

    if SOList.objects.filter(
        user_id=user_id, name__iexact=new_name, deleted_at__isnull=True,
    ).exclude(id=list_id).exists():
        return JsonResponse({'status': 'error', 'message': f'A list named "{new_name}" already exists.'}, status=400)

    lst.name = new_name
    update_fields = ['name']

    if 'tags' in body:
        lst.tags = (body.get('tags') or '').strip()[:255]
        update_fields.append('tags')

    new_status = (body.get('status') or '').strip()
    if new_status in ('active', 'inactive'):
        lst.status = new_status
        update_fields.append('status')

    lst.save(update_fields=update_fields)
    return JsonResponse({
        'status': 'ok', 'name': lst.name, 'tags': lst.tags, 'list_status': lst.status,
    })


def so_list_toggle_status(request, list_id):
    r = _auth_json(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    user_id = get_user_id(request)
    lst = _get_list(user_id, list_id)
    if not lst:
        return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)

    lst.status = 'inactive' if lst.status == 'active' else 'active'
    lst.save(update_fields=['status'])
    return JsonResponse({'status': 'ok', 'new_status': lst.status})


def so_list_duplicate(request, list_id):
    """Copy a list — name, tags and its prospect memberships."""
    r = _auth_json(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    from Email_validate_app.models import SOList, SOListProspect
    user_id = get_user_id(request)
    lst = _get_list(user_id, list_id)
    if not lst:
        return JsonResponse({'status': 'error', 'message': 'List not found.'}, status=404)

    base, name, counter = f'Copy of {lst.name}'[:255], f'Copy of {lst.name}'[:255], 2
    while SOList.objects.filter(user_id=user_id, name__iexact=name, deleted_at__isnull=True).exists():
        name = f'{base} ({counter})'[:255]
        counter += 1

    copy = SOList.objects.create(user_id=user_id, name=name, tags=lst.tags or '', status=lst.status)
    SOListProspect.objects.bulk_create(
        [
            SOListProspect(so_list=copy, prospect_id=pid)
            for pid in SOListProspect.objects.filter(
                so_list=lst, prospect__deleted_at__isnull=True,
            ).values_list('prospect_id', flat=True)
        ],
        ignore_conflicts=True,
    )
    _sync_so_list_counts(copy.id)
    copy.refresh_from_db()

    return JsonResponse({
        'status': 'ok', 'id': copy.id, 'name': copy.name, 'total_count': copy.total_count,
    })


def so_list_download(request, list_id):
    r = _auth(request)
    if r:
        return r

    from Email_validate_app.models import SOListProspect
    user_id = get_user_id(request)
    lst = _get_list(user_id, list_id)
    if not lst:
        return HttpResponse('Not found', status=404)

    rows = (
        SOListProspect.objects
        .filter(so_list=lst, prospect__deleted_at__isnull=True)
        .select_related('prospect')
        .order_by('prospect__email')
    )
    safe_name = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in lst.name)[:50]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="so_list_{safe_name}_{list_id}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Email', 'First Name', 'Last Name', 'Company', 'Phone', 'Status', 'Date Added'])
    for lp in rows:
        p = lp.prospect
        writer.writerow([
            p.email, p.first_name or '', p.last_name or '', p.company or '', p.phone or '',
            p.status, lp.added_at.strftime('%Y-%m-%d') if lp.added_at else '',
        ])
    return response


@_require_POST
def so_list_delete(request):
    r = _auth_json(request)
    if r:
        return r

    from Email_validate_app.models import SOList
    user_id = get_user_id(request)
    try:
        list_id = json.loads(request.body).get('list_id')
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    updated = SOList.objects.filter(
        id=list_id, user_id=user_id, deleted_at__isnull=True,
    ).update(deleted_at=_now())
    if not updated:
        return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)
    return JsonResponse({'status': 'ok'})


def so_list_check(request, list_id):
    """Does this list have any prospects? Drives the 'empty list' popup."""
    r = _auth_json(request)
    if r:
        return r

    from Email_validate_app.models import SOListProspect
    user_id = get_user_id(request)
    if not _get_list(user_id, list_id):
        return JsonResponse({'status': 'error', 'message': 'Not found'}, status=404)

    has_prospects = SOListProspect.objects.filter(
        so_list_id=list_id, prospect__deleted_at__isnull=True,
    ).exists()
    return JsonResponse({'has_prospects': has_prospects})


# ── List detail ───────────────────────────────────────────────────────────────

def so_list_detail(request, list_id):
    r = _auth(request)
    if r:
        return r

    from Email_validate_app.models import SOList
    user_id = get_user_id(request)
    lst = get_object_or_404(SOList, id=list_id, user_id=user_id, deleted_at__isnull=True)
    lst.tags_list = [t.strip() for t in lst.tags.split(',') if t.strip()] if lst.tags else []

    return render(request, 'i_SO_List_Detail.html', {
        'lst':                   lst,
        'pf_statuses':           CONTACT_STATUSES,
        'pf_show_status':        True,
        'pf_search_placeholder': 'Search by name, email or company…',
    })


def so_list_prospects_page(request, list_id):
    """AJAX endpoint — filtered, paginated prospects for one list."""
    r = _auth_json(request)
    if r:
        return r

    import traceback as _tb
    from Email_validate_app.models import SOListProspect
    user_id = get_user_id(request)

    if not _get_list(user_id, list_id):
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    try:
        f = extract_filter_params(request)
        per_page = min(
            int(request.GET.get('per_page', getattr(settings, 'FILTER_DEFAULT_PAGE_SIZE', 25))),
            getattr(settings, 'FILTER_MAX_PAGE_SIZE', 100),
        )

        qs_base = SOListProspect.objects.filter(
            so_list_id=list_id, prospect__deleted_at__isnull=True,
        )
        grand_total = timed_count(qs_base, 'so_list_prospects_total', user_id)

        qs = apply_search(
            qs_base, f['search'],
            'prospect__email', 'prospect__first_name', 'prospect__last_name', 'prospect__company',
        )
        qs = apply_status(qs, f['status'], field='prospect__status')
        qs = apply_date_range(qs, f['date_from'], f['date_to'], field='added_at')
        qs = qs.select_related('prospect').order_by('-added_at')

        paginator = Paginator(qs, per_page)
        page_obj  = paginator.get_page(request.GET.get('page', 1))

        start = page_obj.start_index()
        rows  = []
        for i, lp in enumerate(page_obj):
            p = lp.prospect
            rows.append({
                'id':      p.id,
                'row_num': start + i,
                'name':    f'{p.first_name or ""} {p.last_name or ""}'.strip() or '—',
                'email':   p.email,
                'company': p.company or '—',
                'status':  p.status,
                'date':    lp.added_at.strftime('%b %d, %Y') if lp.added_at else '—',
            })

        stats = {
            'total':            grand_total,
            'subscribed':       timed_count(qs_base.filter(prospect__status='subscribed'),       'so_lp_sub',   user_id),
            'never_subscribed': timed_count(qs_base.filter(prospect__status='never_subscribed'), 'so_lp_never', user_id),
            'unsubscribed':     timed_count(qs_base.filter(prospect__status='unsubscribed'),     'so_lp_unsub', user_id),
        }

        return JsonResponse({
            'status':       'ok',
            'prospects':    rows,
            'total':        paginator.count,
            'grand_total':  grand_total,
            'stats':        stats,
            'page':         page_obj.number,
            'num_pages':    paginator.num_pages,
            'has_previous': page_obj.has_previous(),
            'has_next':     page_obj.has_next(),
            'prev_page':    page_obj.number - 1 if page_obj.has_previous() else None,
            'next_page':    page_obj.number + 1 if page_obj.has_next()     else None,
            'start_index':  start,
            'end_index':    page_obj.end_index(),
        })
    except Exception:
        return JsonResponse({'status': 'error', 'message': _tb.format_exc()}, status=500)


@_require_POST
def so_list_add_prospect(request, list_id):
    """Add a single prospect to a list, creating it in the library if needed."""
    r = _auth_json(request)
    if r:
        return r

    from Email_validate_app.models import SOProspect, SOListProspect
    user_id = get_user_id(request)
    lst = _get_list(user_id, list_id)
    if not lst:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    try:
        body = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    email = (body.get('email') or '').strip().lower()
    if not email or not EMAIL_RE.match(email):
        return JsonResponse({'status': 'error', 'message': 'A valid email address is required.'}, status=400)

    # Consent checkbox drives the intended status, exactly like Email Marketing.
    uploaded_status = 'subscribed' if body.get('consent') == '1' else 'never_subscribed'

    prospect = SOProspect.objects.filter(user_id=user_id, email=email).order_by('id').first()
    if prospect and prospect.deleted_at is None and SOListProspect.objects.filter(
        so_list=lst, prospect=prospect,
    ).exists():
        return JsonResponse({'status': 'error', 'message': f'{email} is already in this list.'}, status=400)

    first_name = (body.get('first_name') or '').strip()
    last_name  = (body.get('last_name')  or '').strip()
    company    = (body.get('company')    or '').strip()
    phone      = (body.get('phone')      or '').strip()

    # A tracked unsubscribe always wins, even if the prospect row says otherwise.
    from Email_validate_app.models import SOEvent
    if SOEvent.objects.filter(
        campaign__user_id=user_id, email=email, event_type='unsubscribed',
    ).exists():
        existing_status = 'unsubscribed'
    else:
        # Consent survives a soft-delete — never reset by re-adding the address.
        existing_status = prospect.status if prospect else None

    final_status = _resolve_status(uploaded_status, existing_status)

    if prospect:
        prospect.first_name = first_name or prospect.first_name
        prospect.last_name  = last_name  or prospect.last_name
        prospect.company    = company    or prospect.company
        prospect.phone      = phone      or prospect.phone
        prospect.status     = final_status
        prospect.deleted_at = None
        prospect.save(update_fields=[
            'first_name', 'last_name', 'company', 'phone', 'status', 'deleted_at', 'updated_at',
        ])
    else:
        prospect = SOProspect.objects.create(
            user_id=user_id, email=email, first_name=first_name, last_name=last_name,
            company=company, phone=phone, status=final_status,
        )

    SOListProspect.objects.get_or_create(so_list=lst, prospect=prospect)
    _sync_so_list_counts(lst.id)

    return JsonResponse({'status': 'ok', 'id': prospect.id, 'status_applied': prospect.status})


def so_list_detail_action(request, list_id):
    """Add prospects from the library / remove prospects from this list."""
    r = _auth_json(request)
    if r:
        return r
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=405)

    from Email_validate_app.models import SOProspect, SOListProspect
    user_id = get_user_id(request)
    lst = _get_list(user_id, list_id)
    if not lst:
        return JsonResponse({'status': 'error', 'message': 'List not found.'}, status=404)

    try:
        data = json.loads(request.body)
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON'}, status=400)

    action = data.get('action')
    ids    = data.get('prospect_ids') or data.get('ids') or []

    if action == 'add_prospects':
        prospects = SOProspect.objects.filter(id__in=ids, user_id=user_id, deleted_at__isnull=True)
        already = set(
            SOListProspect.objects.filter(so_list=lst, prospect__in=prospects)
            .values_list('prospect_id', flat=True)
        )
        to_add = [p for p in prospects if p.id not in already]
        SOListProspect.objects.bulk_create(
            [SOListProspect(so_list=lst, prospect=p) for p in to_add],
            ignore_conflicts=True,
        )
        _sync_so_list_counts(lst.id)
        lst.refresh_from_db()
        return JsonResponse({
            'status': 'ok', 'added': len(to_add), 'skipped': len(already),
            'total_count': lst.total_count,
        })

    if action == 'remove_prospects':
        SOListProspect.objects.filter(so_list=lst, prospect_id__in=ids).delete()
        _sync_so_list_counts(lst.id)
        lst.refresh_from_db()
        return JsonResponse({'status': 'ok', 'removed': len(ids), 'total_count': lst.total_count})

    return JsonResponse({'status': 'error', 'message': 'Unknown action.'})


def so_prospect_detail(request, prospect_id):
    """Full detail JSON for one prospect."""
    r = _auth_json(request)
    if r:
        return r

    from Email_validate_app.models import SOProspect, SOListProspect
    user_id = get_user_id(request)

    try:
        p = SOProspect.objects.get(id=prospect_id, user_id=user_id, deleted_at__isnull=True)
    except SOProspect.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Prospect not found.'}, status=404)

    lists = list(
        SOListProspect.objects.filter(prospect=p, so_list__deleted_at__isnull=True)
        .values_list('so_list__name', flat=True)
    )
    extra = p.extra_data or {}

    return JsonResponse({
        'status': 'ok',
        'prospect': {
            'id':         p.id,
            'first_name': p.first_name,
            'last_name':  p.last_name,
            'email':      p.email,
            'company':    p.company,
            'phone':      p.phone,
            'status':     p.status,
            'lists':      lists,
            'extra_data': extra,
            'created_at': p.created_at.strftime('%d %b %Y, %H:%M') if p.created_at else '—',
            'updated_at': p.updated_at.strftime('%d %b %Y, %H:%M') if p.updated_at else '—',
        },
    })


# ── Upload wizard ─────────────────────────────────────────────────────────────

_NO_HEADER_MSG = (
    'No header row detected. The first row must contain column names '
    '(e.g. "email", "first_name"). Please add a header row and try again.'
)


@_require_POST
def so_list_parse_file(request, list_id):
    """Wizard step 1 — read the uploaded file and return its column names."""
    r = _auth_json(request)
    if r:
        return r

    user_id = get_user_id(request)
    if not _get_list(user_id, list_id):
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    fname = uploaded.name.lower()
    try:
        if fname.endswith('.csv') or fname.endswith('.txt'):
            text = uploaded.read().decode('utf-8-sig', errors='replace')
            if not text.strip():
                return JsonResponse({'status': 'error', 'message': 'The file is empty.'}, status=400)
            try:
                dialect = csv.Sniffer().sniff(text[:2048], delimiters=',\t;|')
            except csv.Error:
                dialect = csv.excel
            reader  = csv.DictReader(io.StringIO(text), dialect=dialect)
            columns = [c.strip() for c in (reader.fieldnames or []) if c and c.strip()]
        elif fname.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            wb.close()
            columns = [str(c).strip() for c in header_row if c is not None and str(c).strip()]
        else:
            return JsonResponse({'status': 'error', 'message': 'Only CSV, XLSX, and TXT files are accepted.'}, status=400)

        if not columns:
            return JsonResponse({'status': 'error', 'message': 'No columns detected. Check the file format.'}, status=400)

        # If the "header" cells look like data (emails / bare numbers), there is no header row
        data_like = sum(
            1 for c in columns
            if EMAIL_RE.match(c) or c.lstrip('-').replace('.', '', 1).isdigit()
        )
        if data_like > 0 and data_like >= len(columns) / 2:
            return JsonResponse({'status': 'error', 'message': _NO_HEADER_MSG}, status=400)

        return JsonResponse({'status': 'ok', 'columns': columns})
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {exc}'}, status=400)


def _read_rows(uploaded):
    """Parse an uploaded CSV/TXT/XLSX into a list of dicts keyed by column name."""
    fname = uploaded.name.lower()
    rows  = []
    if fname.endswith('.csv') or fname.endswith('.txt'):
        text = uploaded.read().decode('utf-8-sig', errors='replace')
        try:
            dialect = csv.Sniffer().sniff(text[:2048], delimiters=',\t;|')
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))
    elif fname.endswith('.xlsx'):
        import openpyxl
        wb     = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
        ws     = wb.active
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c).strip() if c is not None else f'col{j}' for j, c in enumerate(row)]
            else:
                rows.append({
                    header[j]: (str(v).strip() if v is not None else '')
                    for j, v in enumerate(row) if j < len(header)
                })
        wb.close()
    return rows


@_require_POST
def so_list_import_prospects(request, list_id):
    """Wizard step 2 — import the file into the list using the column mapping."""
    r = _auth_json(request)
    if r:
        return r

    from Email_validate_app.models import SOProspect, SOListProspect
    user_id = get_user_id(request)
    lst = _get_list(user_id, list_id)
    if not lst:
        return JsonResponse({'status': 'error', 'message': 'List not found'}, status=404)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)

    try:
        mapping = json.loads(request.POST.get('mapping', '{}'))  # {file_col: system_field}
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Invalid mapping data.'}, status=400)

    if 'email' not in mapping.values():
        return JsonResponse({'status': 'error', 'message': 'Email field must be mapped.'}, status=400)
    if 'first_name' not in mapping.values() and 'last_name' not in mapping.values():
        return JsonResponse({
            'status': 'error',
            'message': 'At least one name field (first_name or last_name) must be mapped.',
        }, status=400)

    email_col      = next(c for c, s in mapping.items() if s == 'email')
    status_col     = next((c for c, s in mapping.items() if s == 'status'),     None)
    first_name_col = next((c for c, s in mapping.items() if s == 'first_name'), None)
    last_name_col  = next((c for c, s in mapping.items() if s == 'last_name'),  None)
    company_col    = next((c for c, s in mapping.items() if s == 'company'),    None)
    phone_col      = next((c for c, s in mapping.items() if s == 'phone'),      None)
    mapped_cols    = set(mapping.keys())

    try:
        rows = _read_rows(uploaded)
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': f'Could not read file: {exc}'}, status=400)

    # ── Sanity-check the mapping against a sample of real rows ────────────────
    sample = [r for r in rows if (r.get(email_col) or '').strip()][:10]
    if sample:
        hits = sum(1 for r in sample if EMAIL_RE.match((r.get(email_col) or '').strip().lower()))
        if hits < max(1, len(sample) * 0.6):
            return JsonResponse({
                'status': 'error',
                'message': (
                    f'The column mapped to Email ("{email_col}") does not appear to contain email '
                    f'addresses (checked {len(sample)} rows, only {hits} looked valid). '
                    f'Please check your column mapping.'
                ),
            }, status=400)
        for label, col in [('First Name', first_name_col), ('Last Name', last_name_col)]:
            if not col:
                continue
            email_like = sum(1 for r in sample if EMAIL_RE.match((r.get(col) or '').strip().lower()))
            if email_like >= len(sample) * 0.3:
                return JsonResponse({
                    'status': 'error',
                    'message': (
                        f'The column mapped to {label} ("{col}") appears to contain email addresses '
                        f'rather than names. Please check your column mapping.'
                    ),
                }, status=400)

    # ── Existing state: library statuses + who is already in this list ────────
    from Email_validate_app.models import SOEvent

    existing_prospects = {
        p['email']: p for p in SOProspect.objects.filter(user_id=user_id)
        .values('id', 'email', 'status', 'deleted_at')
    }
    in_list_emails = set(
        SOListProspect.objects.filter(so_list=lst, prospect__deleted_at__isnull=True)
        .values_list('prospect__email', flat=True)
    )
    # Tracked unsubscribes always win and can never be re-subscribed.
    hard_unsub = set(
        SOEvent.objects.filter(campaign__user_id=user_id, event_type='unsubscribed')
        .values_list('email', flat=True)
    )

    def existing_status(email):
        # Deliberately ignores deleted_at — an unsubscribe is a consent signal
        # that survives a soft-delete of the prospect record.
        if email in hard_unsub:
            return 'unsubscribed'
        p = existing_prospects.get(email)
        return p['status'] if p else None

    # ── Pre-flight: rows whose status will be overridden by library state ─────
    if not request.POST.get('confirmed'):
        conflicts = []
        for i, row in enumerate(rows):
            email = (row.get(email_col) or '').strip().lower()
            if not email or not EMAIL_RE.match(email) or email in in_list_emails:
                continue
            uploaded_status = _normalize_status(row.get(status_col)) if status_col else 'subscribed'
            if uploaded_status is None:
                continue
            existing = existing_status(email)
            if existing is None:
                continue
            final = _resolve_status(uploaded_status, existing)
            if final != uploaded_status:
                conflicts.append({
                    'row':      i + 2,
                    'email':    email,
                    'uploaded': uploaded_status,
                    'existing': existing,
                    'final':    final,
                })
        if conflicts:
            return JsonResponse({
                'status':          'conflicts',
                'conflicts':       conflicts[:100],
                'total_conflicts': len(conflicts),
            })

    # ── Import ────────────────────────────────────────────────────────────────
    total        = len(rows)
    seen         = set()
    skipped_rows = []
    to_create    = []
    to_update    = []
    link_ids     = []
    by_status    = {'subscribed': 0, 'unsubscribed': 0, 'never_subscribed': 0}

    for i, row in enumerate(rows):
        raw_email  = (row.get(email_col) or '').strip().lower()
        raw_status = (row.get(status_col) or '').strip() if status_col else ''

        if not raw_email:
            skipped_rows.append({'row': i + 2, 'email': '(empty)', 'uploaded_status': raw_status, 'reason': 'Missing email'})
            continue
        if not EMAIL_RE.match(raw_email):
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status, 'reason': 'Invalid email format'})
            continue
        if raw_email in seen:
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status, 'reason': 'Duplicate row in file'})
            continue
        if raw_email in in_list_emails:
            skipped_rows.append({'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status, 'reason': 'Already exists in this list'})
            continue

        if status_col:
            uploaded_status = _normalize_status(raw_status)
            if uploaded_status is None:
                skipped_rows.append({
                    'row': i + 2, 'email': raw_email, 'uploaded_status': raw_status,
                    'reason': f'Invalid status value "{raw_status}"',
                })
                continue
        else:
            uploaded_status = 'subscribed'

        seen.add(raw_email)
        final_status = _resolve_status(uploaded_status, existing_status(raw_email))
        by_status[final_status] = by_status.get(final_status, 0) + 1

        first = (row.get(first_name_col) or '').strip() if first_name_col else ''
        last  = (row.get(last_name_col)  or '').strip() if last_name_col  else ''
        comp  = (row.get(company_col)    or '').strip() if company_col    else ''
        phone = (row.get(phone_col)      or '').strip() if phone_col      else ''
        extra = {
            col: (row.get(col) or '').strip()
            for col in row if col not in mapped_cols and (row.get(col) or '').strip()
        }

        known = existing_prospects.get(raw_email)
        if known:
            to_update.append((known['id'], first, last, comp, phone, final_status, extra))
            link_ids.append(known['id'])
        else:
            to_create.append(SOProspect(
                user_id=user_id, email=raw_email, first_name=first, last_name=last,
                company=comp, phone=phone, status=final_status, extra_data=extra,
            ))

    if to_create:
        SOProspect.objects.bulk_create(to_create, ignore_conflicts=True)
        link_ids += list(
            SOProspect.objects.filter(user_id=user_id, email__in=[p.email for p in to_create])
            .values_list('id', flat=True)
        )

    for pid, first, last, comp, phone, final_status, extra in to_update:
        p = SOProspect.objects.filter(id=pid).first()
        if not p:
            continue
        p.first_name = first or p.first_name
        p.last_name  = last  or p.last_name
        p.company    = comp  or p.company
        p.phone      = phone or p.phone
        p.status     = final_status
        p.deleted_at = None
        if extra:
            merged = dict(p.extra_data or {})
            merged.update(extra)
            p.extra_data = merged
        p.save()

    SOListProspect.objects.bulk_create(
        [SOListProspect(so_list=lst, prospect_id=pid) for pid in set(link_ids)],
        ignore_conflicts=True,
    )
    _sync_so_list_counts(lst.id)

    imported = len(seen)
    return JsonResponse({
        'status': 'ok',
        'summary': {
            'total':        total,
            'imported':     imported,
            'by_status':    by_status,
            'skipped':      len(skipped_rows),
            'skipped_rows': skipped_rows[:500],
        },
    })
